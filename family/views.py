from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.http import FileResponse
from django.contrib import messages
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response

from .forms import FamilyHeadForm, HobbyFormSet, MemberFormset
from .models import FamilyHead, FamilyMember, Hobby, City, statusChoice
from .utils import decode_id

from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image as ExcelImage

import os, logging, json

logger = logging.getLogger(__name__)


def home(request):
    try:
        return render(request, 'index.html')
    except Exception as e:
        logger.exception("Error loading home page: %s", e)
        return HttpResponse("An unexpected error occurred.", status=500)


@api_view(['GET'])
@permission_classes([AllowAny])
def get_cities(request, pk):
    try:
        cities = City.objects.filter(state_id=pk, status=statusChoice.ACTIVE)
        data = list(cities.values('id', 'city_name'))
        return Response(data)
    except Exception as e:
        logger.exception("Error fetching cities: %s", e)
        return Response({"error": "Unable to load cities."}, status=500)


def family_form(request):
    try:
        head_form = FamilyHeadForm()
        hobby_formset = HobbyFormSet(prefix="hobbies")
        member_formset = MemberFormset(prefix="members")

        if request.method == 'POST':
            head_form = FamilyHeadForm(request.POST, request.FILES)
            hobby_formset = HobbyFormSet(request.POST, instance=head_form.instance, prefix="hobbies")
            member_formset = MemberFormset(request.POST, request.FILES, instance=head_form.instance, prefix="members")

            if head_form.is_valid() and hobby_formset.is_valid() and member_formset.is_valid():
                head = head_form.save()
                hobby_formset.instance = head
                hobby_formset.save()
                member_formset.instance = head
                member_formset.save()
                return JsonResponse({"success": True, "message": "Family Created Successfully."})
            else:
                return JsonResponse({
                    "success": False,
                    "head_errors": head_form.errors,
                    "hobby_errors": hobby_formset.errors,
                    "member_errors": member_formset.errors,
                }, status=400)

        context = {'head_form': head_form, 'hobby_formset': hobby_formset, 'member_formset': member_formset}
        return render(request, 'family_form.html', context)

    except Exception as e:
        logger.exception("Error in family_form: %s", e)
        return JsonResponse({"success": False, "errorMessage": "Unexpected error occurred while saving family."}, status=500)


@login_required(login_url='login_page')
def family_pdf(request, hashid):
    try:
        pk = decode_id(hashid)
        head = FamilyHead.objects.get(pk=pk)
        members = FamilyMember.objects.filter(family_head=head, status=statusChoice.ACTIVE)
        hobbies = Hobby.objects.filter(family_head=head, status=statusChoice.ACTIVE)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{head.name}_family.pdf"'

        doc = SimpleDocTemplate(response)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(f"Family Report: {head.surname} Family", styles['Heading1']))
        elements.append(Paragraph("Head Details", styles['Heading2']))
        elements.append(Spacer(1, 4))

        details = [
            f"Name: {head.name}",
            f"Surname: {head.surname}",
            f"Birth Date: {head.dob}",
            f"Mobile: {head.mobno}",
            f"Address: {head.address}",
            f"State: {head.state.state_name}",
            f"City: {head.city.city_name}",
            f"Pincode: {head.pincode}",
            f"Marital Status: {head.marital_status}",
            f"Wedding Date: {head.wedding_date}",
        ]
        for d in details:
            elements.append(Paragraph(d, styles['Normal']))
            elements.append(Spacer(1, 4))

        # Head Photo
        if head.photo and hasattr(head.photo, 'path') and os.path.exists(head.photo.path):
            try:
                img = Image(head.photo.path, width=1.5 * inch, height=2 * inch)
                img.hAlign = 'CENTER'
                elements.append(Paragraph("Photo:", styles['Normal']))
                elements.append(img)
                elements.append(Spacer(1, 12))
            except Exception as img_error:
                logger.warning("Error adding head photo: %s", img_error)

        # Hobbies
        elements.append(Paragraph("Hobbies", styles['Heading3']))
        for i, h in enumerate(hobbies, start=1):
            elements.append(Paragraph(f"{i}. {h.hobby}", styles['Normal']))
            elements.append(Spacer(1, 4))

        # Members
        elements.append(Paragraph("Members", styles['Heading2']))
        for i, m in enumerate(members, start=1):
            elements.append(Paragraph(f"Member {i}", styles['Heading3']))
            member_details = [
                f"Name: {m.member_name}",
                f"Birth Date: {m.member_dob}",
                f"Marital Status: {m.member_marital}",
                f"Wedding Date: {m.member_wedDate}",
                f"Education: {m.education}",
                f"Relation: {m.relation}",
            ]
            for d in member_details:
                elements.append(Paragraph(d, styles['Normal']))
                elements.append(Spacer(1, 4))

            if m.member_photo and hasattr(m.member_photo, 'path') and os.path.exists(m.member_photo.path):
                try:
                    img = Image(m.member_photo.path, width=1.5 * inch, height=2 * inch)
                    img.hAlign = 'CENTER'
                    elements.append(img)
                    elements.append(Spacer(1, 12))
                except Exception as img_error:
                    logger.warning("Error adding member photo: %s", img_error)

        doc.build(elements)
        return response

    except FamilyHead.DoesNotExist:
        messages.error(request, "Family not found.")
        return redirect('dashboard')
    except Exception as e:
        logger.exception("Error generating family PDF: %s", e)
        messages.error(request, "Error while generating PDF. Please try again.")
        return redirect('dashboard')


@login_required(login_url='login_page')
def family_excel(request, hashid):
    try:
        pk = decode_id(hashid)
        head = FamilyHead.objects.get(id=pk)
        hobbies = Hobby.objects.filter(family_head=pk, status=statusChoice.ACTIVE)
        members = FamilyMember.objects.filter(family_head=pk, status=statusChoice.ACTIVE)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{head.name}_family.xlsx"'
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Family Report'

        worksheet.merge_cells('A1:L1')
        worksheet['A1'].value = "Family Report"
        worksheet['A1'].fill = PatternFill("solid", fgColor="246ba1")
        worksheet['A1'].font = Font(bold=True, color="F7F6FA")
        worksheet['A1'].alignment = Alignment(horizontal="center")

        columns = ['Name', 'Surname', 'Birth Date', 'Mobile No', 'Address', 'State', 'City',
                   'Pincode', 'Marital Status', 'Wedding Date', 'Photo', 'Hobbies']
        worksheet.append(columns)

        hobbies_str = ", ".join([h.hobby for h in hobbies])
        worksheet.append([
            head.name, head.surname, str(head.dob), head.mobno, head.address,
            head.state.state_name, head.city.city_name, head.pincode,
            head.marital_status, str(head.wedding_date), str(head.photo), hobbies_str
        ])

        # Add head image
        if head.photo and hasattr(head.photo, 'path') and os.path.exists(head.photo.path):
            try:
                img = ExcelImage(head.photo.path)
                img.width, img.height = 50, 50
                worksheet.add_image(img, 'K3')
            except Exception as img_error:
                logger.warning("Error adding head image in Excel: %s", img_error)

        # Add member rows
        worksheet.append(['', 'Member Details'])
        worksheet.append(['Sr. No.', 'Name', 'Birth Date', 'Marital Status', 'Wedding Date', 'Education', 'Photo'])
        for i, m in enumerate(members, start=1):
            worksheet.append([
                i, m.member_name, str(m.member_dob), m.member_marital,
                str(m.member_wedDate), m.education, str(m.member_photo)
            ])
            if m.member_photo and hasattr(m.member_photo, 'path') and os.path.exists(m.member_photo.path):
                try:
                    img = ExcelImage(m.member_photo.path)
                    img.width, img.height = 50, 50
                    worksheet.add_image(img, f'G{worksheet.max_row}')
                except Exception as img_error:
                    logger.warning("Error adding member image in Excel: %s", img_error)

        workbook.save(response)
        return response

    except FamilyHead.DoesNotExist:
        messages.error(request, "Family not found.")
        return redirect('dashboard')
    except Exception as e:
        logger.exception("Error generating family Excel: %s", e)
        messages.error(request, "Error while exporting to Excel.")
        return redirect('dashboard')


@login_required(login_url='login_page')
def head_excel(request):
    try:
        heads = FamilyHead.objects.exclude(status=statusChoice.DELETE)

        # Filtering by search
        search = request.GET.get('search')
        if search:
            name = heads.filter(name__icontains=search)
            mobno = heads.filter(mobno__icontains=search)
            state = heads.filter(state__state_name__icontains=search)
            city = heads.filter(city__city_name__icontains=search)
            heads = name.union(mobno, state, city)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="all_family_heads.xlsx"'
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'All Family Head Report'

        worksheet.merge_cells('A1:Q1')
        worksheet['A1'].value = "All Family Head Report"
        worksheet['A1'].fill = PatternFill("solid", fgColor="246ba1")
        worksheet['A1'].font = Font(bold=True, color="F7F6FA")
        worksheet['A1'].alignment = Alignment(horizontal="center")

        columns = [
            'Sr. No.', 'Member ID', 'Name', 'Surname', 'Birth Date', 'Mobile No',
            'Address', 'State', 'City', 'Pincode', 'Marital Status', 'Wedding Date',
            'Education', 'Relation', 'Photo', 'Hobbies', 'Head ID'
        ]
        worksheet.append(columns)

        for i, head in enumerate(heads, start=1):
            hobbies = Hobby.objects.filter(family_head=head.id, status=statusChoice.ACTIVE)
            hobbies_str = ", ".join([h.hobby for h in hobbies])

            worksheet.append([
                i, "", head.name, head.surname, str(head.dob), head.mobno, head.address,
                head.state.state_name, head.city.city_name, head.pincode, head.marital_status,
                str(head.wedding_date), "", "Head", str(head.photo), hobbies_str, head.id
            ])

            # Head photo
            if head.photo and hasattr(head.photo, 'path') and os.path.exists(head.photo.path):
                try:
                    img = ExcelImage(head.photo.path)
                    img.width, img.height = 30, 30
                    worksheet.add_image(img, f'O{worksheet.max_row}')
                except Exception as img_error:
                    logger.warning("Error adding head image in head_excel: %s", img_error)

            # Members
            members = FamilyMember.objects.filter(family_head=head.id, status=statusChoice.ACTIVE)
            for j, m in enumerate(members, start=1):
                worksheet.append([
                    "", j, m.member_name, "", str(m.member_dob), "-", "", "", "", "",
                    m.member_marital, str(m.member_wedDate), m.education,
                    m.relation, str(m.member_photo), "", m.family_head.id
                ])
                if m.member_photo and hasattr(m.member_photo, 'path') and os.path.exists(m.member_photo.path):
                    try:
                        img = ExcelImage(m.member_photo.path)
                        img.width, img.height = 30, 30
                        worksheet.add_image(img, f'O{worksheet.max_row}')
                    except Exception as img_error:
                        logger.warning("Error adding member image in head_excel: %s", img_error)

        workbook.save(response)
        return response

    except Exception as e:
        logger.exception("Error generating head Excel report: %s", e)
        messages.error(request, "Error exporting family head data.")
        return redirect('dashboard')
