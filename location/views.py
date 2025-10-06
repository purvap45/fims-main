from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from family.models import State, City, statusChoice
from .forms import StateForm, CityForm
from family.utils import decode_id

# ----------------------------- STATE VIEWS -----------------------------

@login_required(login_url='login_page')
def state_list(request):
    try:
        states = State.objects.exclude(status=statusChoice.DELETE).order_by('-created_at')

        search = request.GET.get('search')
        if search:
            states = states.filter(Q(state_name__icontains=search))

        paginator = Paginator(states, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        total_pages = page_obj.paginator.num_pages

        context = {
            'page_obj': page_obj,
            'lastPage': total_pages,
            'totalPagelist': [n + 1 for n in range(total_pages)],
        }

        # Handle AJAX pagination/search
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            state_html = render_to_string('state_list_template.html', context, request=request)
            return JsonResponse({'state_html': state_html})

        return render(request, 'state_list.html', context)

    except Exception as e:
        messages.error(request, f"An error occurred while loading states: {str(e)}")
        return redirect('dashboard')


@login_required(login_url='login_page')
def create_state(request):
    try:
        state_form = StateForm(request.POST or None)
        if request.method == 'POST':
            if state_form.is_valid():
                state_form.save()
                messages.success(request, 'State created successfully!')
                return redirect('state_list')
            else:
                messages.error(request, 'Please correct the errors below.')
        return render(request, 'create_state.html', {'state_form': state_form})
    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {str(e)}")
        return redirect('state_list')


@login_required(login_url='login_page')
def update_state(request, hashid):
    try:
        pk = decode_id(hashid)
        state = get_object_or_404(State, id=pk)

        state_form = StateForm(request.POST or None, instance=state)
        if request.method == 'POST' and state_form.is_valid():
            state = state_form.save()
            state.set_status(state.status)
            messages.success(request, 'State updated successfully!')
            return redirect('state_list')

        return render(request, 'update_state.html', {'state_form': state_form})

    except ValueError:
        messages.error(request, 'Invalid state ID.')
    except State.DoesNotExist:
        messages.error(request, 'State not found.')
    except Exception as e:
        messages.error(request, f"Error updating state: {str(e)}")
    return redirect('state_list')


@login_required(login_url='login_page')
def delete_state(request, hashid):
    try:
        pk = decode_id(hashid)
        state = get_object_or_404(State, id=pk)
        City.objects.filter(state=state).update(status=statusChoice.DELETE)
        state.soft_delete()
        messages.success(request, 'State and related cities deleted successfully!')
    except ValueError:
        messages.error(request, 'Invalid state ID.')
    except Exception as e:
        messages.error(request, f"Error deleting state: {str(e)}")
    return redirect('state_list')


@login_required(login_url='login_page')
def state_excel(request):
    try:
        states = State.objects.exclude(status=statusChoice.DELETE)
        search = request.GET.get('search')
        if search:
            states = states.filter(state_name__icontains=search)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="state.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'State'

        worksheet.merge_cells('A1:C1')
        worksheet['A1'].value = "State List"
        worksheet['A1'].fill = PatternFill("solid", fgColor="246BA1")
        worksheet['A1'].font = Font(bold=True, color="FFFFFF")
        worksheet['A1'].alignment = Alignment(horizontal="center")

        worksheet.append(['ID', 'Name', 'Status'])

        for count, state in enumerate(states, start=1):
            worksheet.append([count, state.state_name, state.status])

        workbook.save(response)
        return response

    except Exception as e:
        messages.error(request, f"Error exporting states: {str(e)}")
        return redirect('state_list')


# ----------------------------- CITY VIEWS -----------------------------

@login_required(login_url='login_page')
def city_list(request):
    try:
        cities = City.objects.exclude(status=statusChoice.DELETE).order_by('-created_at')
        search = request.GET.get('search')
        if search:
            cities = cities.filter(
                Q(city_name__icontains=search) | Q(state__state_name__icontains=search)
            )

        paginator = Paginator(cities, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        total_pages = page_obj.paginator.num_pages

        context = {
            'page_obj': page_obj,
            'lastPage': total_pages,
            'totalPagelist': [n + 1 for n in range(total_pages)],
        }

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            city_html = render_to_string('city_list_template.html', context, request=request)
            return JsonResponse({'city_html': city_html})

        return render(request, 'city_list.html', context)

    except Exception as e:
        messages.error(request, f"An error occurred while loading cities: {str(e)}")
        return redirect('dashboard')


@login_required(login_url='login_page')
def create_city(request):
    try:
        city_form = CityForm(request.POST or None)
        if request.method == 'POST':
            if city_form.is_valid():
                city_form.save()
                messages.success(request, 'City created successfully!')
                return redirect('city_list')
            else:
                messages.error(request, 'Please correct the errors below.')
        return render(request, 'create_city.html', {'city_form': city_form})
    except Exception as e:
        messages.error(request, f"An unexpected error occurred: {str(e)}")
        return redirect('city_list')


@login_required(login_url='login_page')
def update_city(request, hashid):
    try:
        pk = decode_id(hashid)
        city = get_object_or_404(City, id=pk)

        city_form = CityForm(request.POST or None, instance=city)
        if request.method == 'POST' and city_form.is_valid():
            city = city_form.save()
            city.set_status(city.status)
            messages.success(request, 'City updated successfully!')
            return redirect('city_list')

        return render(request, 'update_city.html', {'city_form': city_form})

    except ValueError:
        messages.error(request, 'Invalid city ID.')
    except City.DoesNotExist:
        messages.error(request, 'City not found.')
    except Exception as e:
        messages.error(request, f"Error updating city: {str(e)}")
    return redirect('city_list')


@login_required(login_url='login_page')
def delete_city(request, hashid):
    try:
        pk = decode_id(hashid)
        city = get_object_or_404(City, id=pk)
        city.soft_delete()
        messages.success(request, 'City deleted successfully!')
    except ValueError:
        messages.error(request, 'Invalid city ID.')
    except Exception as e:
        messages.error(request, f"Error deleting city: {str(e)}")
    return redirect('city_list')


@login_required(login_url='login_page')
def city_excel(request):
    try:
        cities = City.objects.exclude(status=statusChoice.DELETE)
        search = request.GET.get('search')
        if search:
            cities = cities.filter(
                Q(city_name__icontains=search) | Q(state__state_name__icontains=search)
            )

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="city.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'City'

        worksheet.merge_cells('A1:D1')
        worksheet['A1'].value = "City List"
        worksheet['A1'].fill = PatternFill("solid", fgColor="246BA1")
        worksheet['A1'].font = Font(bold=True, color="FFFFFF")
        worksheet['A1'].alignment = Alignment(horizontal="center")

        worksheet.append(['ID', 'Name', 'State', 'Status'])

        for count, city in enumerate(cities, start=1):
            worksheet.append([count, city.city_name, city.state.state_name, city.status])

        workbook.save(response)
        return response

    except Exception as e:
        messages.error(request, f"Error exporting cities: {str(e)}")
        return redirect('city_list')
